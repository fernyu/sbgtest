#!/usr/bin/python
# -*- coding: <utf-8> -*-
import os
from util.BrowserDriver import BrowserDriver
import unittest
import openpyxl
from util.BasePage import BasePage
from util.logger import Logger

logger = Logger(logger="DesktopHomepageTest").getlog()


class DesktopHomepageTest(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        driver = BrowserDriver(cls)
        cls.driver = driver.open_browser(cls)

    def setUp(self):
        pass

    def test_homepage(self):
        file_path = os.path.dirname(os.getcwd())
        excel_path = file_path + "/files/website.xlsx"
        sheet_name = "Sheet1"
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb[sheet_name]
        for i in range(1, sheet.max_row + 1, 1):
            web_url = sheet.cell(row=i, column=1).value
            try:
                if web_url is None:
                    break
                else:
                    self.driver.get(web_url)
                    title = self.driver.title
                    if "Error" in title:
                        logger.info(web_url + " " + "ERROR")
                        BasePage().get_screent_img()
                    elif title == web_url[7:]:
                        logger.info(web_url + " " + "ERROR")
                    else:
                        logger.info(web_url + " " + "homepage is OK")
            except Exception as e:
                logger.info(e)

    def tearDown(self):
        pass

    @classmethod
    def tearDownClass(cls):
        cls.driver.quit()


if __name__ == "__main__":
    unittest.main()
